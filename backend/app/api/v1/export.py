"""
数据导出 API

支持的导出类型：
- contributions: 贡献记录
- rewards: 回报记录
- actors: Actor数据
- tasks: 任务数据

支持的导出格式：
- csv: CSV格式
- excel: Excel格式
- json: JSON格式
"""

import io
import csv
import json
from datetime import datetime, timedelta
from typing import Optional, List
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field

from app.database import get_db
from app.models.contribution import (
    Actor, Event, Reward, Task,
    ContributionStatus, RewardStatus, ActorType
)


router = APIRouter(prefix="/export", tags=["数据导出"])


# ========== 导出请求模型 ==========

class ExportRequest(BaseModel):
    """导出请求"""
    data_type: str = Field(..., description="数据类型: contributions/rewards/actors/tasks")
    format: str = Field(default="csv", description="导出格式: csv/excel/json")
    start_date: Optional[str] = Field(None, description="开始日期 (YYYY-MM-DD)")
    end_date: Optional[str] = Field(None, description="结束日期 (YYYY-MM-DD)")
    status: Optional[str] = Field(None, description="状态筛选")
    actor_id: Optional[str] = Field(None, description="Actor ID筛选")


# ========== 辅助函数 ==========

def parse_date(date_str: Optional[str]) -> Optional[datetime]:
    """解析日期字符串"""
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return None


def generate_csv(data: List[dict], headers: List[str]) -> str:
    """生成CSV内容"""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction='ignore')
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue()


def generate_excel(data: List[dict], headers: List[str], sheet_name: str = "Sheet1") -> bytes:
    """生成Excel内容"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel导出需要安装openpyxl库: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.border = thin_border
    
    # 自动调整列宽
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row_data in data:
            cell_value = str(row_data.get(header, ""))
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def generate_json(data: List[dict]) -> str:
    """生成JSON内容"""
    return json.dumps(data, ensure_ascii=False, indent=2, default=str)


def get_export_response(
    content: bytes | str,
    format: str,
    filename: str
) -> Response:
    """生成导出响应"""
    media_types = {
        "csv": "text/csv",
        "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "json": "application/json"
    }
    
    extensions = {
        "csv": ".csv",
        "excel": ".xlsx",
        "json": ".json"
    }
    
    if isinstance(content, str):
        content = content.encode('utf-8')
    
    return Response(
        content=content,
        media_type=media_types.get(format, "application/octet-stream"),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}{extensions.get(format, ".txt")}"'
        }
    )


# ========== 导出端点 ==========

@router.post("", summary="导出数据")
async def export_data(
    request: ExportRequest,
    db: Session = Depends(get_db)
):
    """
    导出数据
    
    - data_type: contributions/rewards/actors/tasks
    - format: csv/excel/json
    - start_date/end_date: 时间范围
    - status: 状态筛选
    - actor_id: Actor ID筛选
    """
    data = []
    headers = []
    filename = f"{request.data_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    # 解析日期
    start_date = parse_date(request.start_date)
    end_date = parse_date(request.end_date)
    if end_date:
        end_date = end_date.replace(hour=23, minute=59, second=59)
    
    # 根据数据类型获取数据
    if request.data_type == "contributions":
        data, headers = export_contributions(db, start_date, end_date, request.status, request.actor_id)
        filename = f"贡献记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    elif request.data_type == "rewards":
        data, headers = export_rewards(db, start_date, end_date, request.status, request.actor_id)
        filename = f"回报记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    elif request.data_type == "actors":
        data, headers = export_actors(db, request.status)
        filename = f"Actor数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    elif request.data_type == "tasks":
        data, headers = export_tasks(db, start_date, end_date, request.status)
        filename = f"任务数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    else:
        raise HTTPException(status_code=400, detail=f"不支持的数据类型: {request.data_type}")
    
    if not data:
        raise HTTPException(status_code=404, detail="没有符合条件的数据可导出")
    
    # 根据格式生成内容
    if request.format == "csv":
        content = generate_csv(data, headers)
    elif request.format == "excel":
        content = generate_excel(data, headers, request.data_type)
    elif request.format == "json":
        content = generate_json(data)
    else:
        raise HTTPException(status_code=400, detail=f"不支持的导出格式: {request.format}")
    
    return get_export_response(content, request.format, filename)


@router.get("/contributions", summary="导出贡献记录")
async def export_contributions_get(
    format: str = Query("csv", description="导出格式: csv/excel/json"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="状态筛选"),
    actor_id: Optional[str] = Query(None, description="Actor ID筛选"),
    db: Session = Depends(get_db)
):
    """导出贡献记录 (GET方式)"""
    request = ExportRequest(
        data_type="contributions",
        format=format,
        start_date=start_date,
        end_date=end_date,
        status=status,
        actor_id=actor_id
    )
    return await export_data(request, db)


@router.get("/rewards", summary="导出回报记录")
async def export_rewards_get(
    format: str = Query("csv", description="导出格式: csv/excel/json"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="状态筛选"),
    actor_id: Optional[str] = Query(None, description="Actor ID筛选"),
    db: Session = Depends(get_db)
):
    """导出回报记录 (GET方式)"""
    request = ExportRequest(
        data_type="rewards",
        format=format,
        start_date=start_date,
        end_date=end_date,
        status=status,
        actor_id=actor_id
    )
    return await export_data(request, db)


@router.get("/actors", summary="导出Actor数据")
async def export_actors_get(
    format: str = Query("csv", description="导出格式: csv/excel/json"),
    actor_type: Optional[str] = Query(None, description="Actor类型: human/ai_agent/external"),
    db: Session = Depends(get_db)
):
    """导出Actor数据 (GET方式)"""
    request = ExportRequest(
        data_type="actors",
        format=format,
        status=actor_type
    )
    return await export_data(request, db)


@router.get("/tasks", summary="导出任务数据")
async def export_tasks_get(
    format: str = Query("csv", description="导出格式: csv/excel/json"),
    start_date: Optional[str] = Query(None, description="开始日期 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="结束日期 (YYYY-MM-DD)"),
    status: Optional[str] = Query(None, description="状态筛选"),
    db: Session = Depends(get_db)
):
    """导出任务数据 (GET方式)"""
    request = ExportRequest(
        data_type="tasks",
        format=format,
        start_date=start_date,
        end_date=end_date,
        status=status
    )
    return await export_data(request, db)


# ========== 数据获取函数 ==========

def export_contributions(
    db: Session,
    start_date: Optional[datetime],
    end_date: Optional[datetime],
    status: Optional[str],
    actor_id: Optional[str]
) -> tuple[List[dict], List[str]]:
    """导出贡献记录"""
    query = db.query(Event)
    
    if start_date:
        query = query.filter(Event.created_at >= start_date)
    if end_date:
        query = query.filter(Event.created_at <= end_date)
    if status:
        query = query.filter(Event.contribution_status == status)
    if actor_id:
        query = query.filter(Event.actor_id == UUID(actor_id))
    
    events = query.order_by(Event.created_at.desc()).all()
    
    # 获取Actor名称映射
    actor_names = {}
    actors = db.query(Actor).all()
    for actor in actors:
        actor_names[str(actor.id)] = actor.display_name
    
    headers = ["ID", "贡献者", "贡献类型", "内容", "预估价值", "实际价值", "置信度", "状态", "创建时间", "验证时间"]
    
    data = []
    for event in events:
        data.append({
            "ID": str(event.id),
            "贡献者": actor_names.get(str(event.actor_id), "未知"),
            "贡献类型": event.contribution_type or "-",
            "内容": (event.content[:100] + "...") if event.content and len(event.content) > 100 else (event.content or "-"),
            "预估价值": float(event.contribution_value) if event.contribution_value else 0,
            "实际价值": float(event.actual_value) if event.actual_value else 0,
            "置信度": float(event.value_confidence) if event.value_confidence else 0,
            "状态": {"pending": "待验证", "verified": "已验证", "rejected": "已拒绝"}.get(event.contribution_status, event.contribution_status),
            "创建时间": event.created_at.strftime("%Y-%m-%d %H:%M:%S") if event.created_at else "-",
            "验证时间": event.verified_at.strftime("%Y-%m-%d %H:%M:%S") if event.verified_at else "-"
        })
    
    return data, headers


def export_rewards(
    db: Session,
    start_date: Optional[datetime],
    end_date: Optional[datetime],
    status: Optional[str],
    actor_id: Optional[str]
) -> tuple[List[dict], List[str]]:
    """导出回报记录"""
    query = db.query(Reward)
    
    if start_date:
        query = query.filter(Reward.created_at >= start_date)
    if end_date:
        query = query.filter(Reward.created_at <= end_date)
    if status:
        query = query.filter(Reward.status == status)
    if actor_id:
        query = query.filter(Reward.actor_id == UUID(actor_id))
    
    rewards = query.order_by(Reward.created_at.desc()).all()
    
    # 获取Actor名称映射
    actor_names = {}
    actors = db.query(Actor).all()
    for actor in actors:
        actor_names[str(actor.id)] = actor.display_name
    
    headers = ["ID", "获得者", "回报类型", "金额", "来源类型", "状态", "创建时间", "兑现时间"]
    
    data = []
    for reward in rewards:
        data.append({
            "ID": str(reward.id),
            "获得者": actor_names.get(str(reward.actor_id), "未知"),
            "回报类型": {"cash_bonus": "现金奖励", "profit_sharing": "利润分红", "stock_options": "期权激励", "skill_points": "技能点"}.get(reward.reward_type, reward.reward_type),
            "金额": float(reward.amount),
            "来源类型": reward.contribution_type or "-",
            "状态": {"pending": "待兑现", "claimed": "已兑现", "expired": "已过期"}.get(reward.status, reward.status),
            "创建时间": reward.created_at.strftime("%Y-%m-%d %H:%M:%S") if reward.created_at else "-",
            "兑现时间": reward.claimed_at.strftime("%Y-%m-%d %H:%M:%S") if reward.claimed_at else "-"
        })
    
    return data, headers


def export_actors(
    db: Session,
    actor_type: Optional[str]
) -> tuple[List[dict], List[str]]:
    """导出Actor数据"""
    query = db.query(Actor)
    
    if actor_type:
        query = query.filter(Actor.actor_type == actor_type)
    
    actors = query.all()
    
    headers = ["ID", "代号", "类型", "总贡献值", "总回报", "可兑现回报", "信任分数", "声誉分数", "能力数量", "贡献次数", "创建时间"]
    
    data = []
    for actor in actors:
        capabilities = actor.capabilities or {}
        data.append({
            "ID": str(actor.id),
            "代号": actor.display_name,
            "类型": {"human": "真实员工", "ai_agent": "AI员工", "external": "外部合作者"}.get(actor.actor_type, actor.actor_type),
            "总贡献值": float(actor.total_contributions) if actor.total_contributions else 0,
            "总回报": float(actor.total_rewards) if actor.total_rewards else 0,
            "可兑现回报": float(actor.available_rewards) if actor.available_rewards else 0,
            "信任分数": float(actor.trust_score) if actor.trust_score else 0.5,
            "声誉分数": float(actor.reputation_score) if actor.reputation_score else 0.5,
            "能力数量": len(capabilities),
            "贡献次数": actor.contribution_count or 0,
            "创建时间": actor.created_at.strftime("%Y-%m-%d %H:%M:%S") if actor.created_at else "-"
        })
    
    return data, headers


def export_tasks(
    db: Session,
    start_date: Optional[datetime],
    end_date: Optional[datetime],
    status: Optional[str]
) -> tuple[List[dict], List[str]]:
    """导出任务数据"""
    query = db.query(Task)
    
    if start_date:
        query = query.filter(Task.created_at >= start_date)
    if end_date:
        query = query.filter(Task.created_at <= end_date)
    if status:
        query = query.filter(Task.status == status)
    
    tasks = query.order_by(Task.created_at.desc()).all()
    
    headers = ["ID", "类型", "描述", "优先级", "状态", "截止时间", "创建时间", "完成时间"]
    
    data = []
    for task in tasks:
        data.append({
            "ID": str(task.id),
            "类型": task.type,
            "描述": (task.description[:100] + "...") if task.description and len(task.description) > 100 else (task.description or "-"),
            "优先级": task.priority or 50,
            "状态": {"pending": "待处理", "in_progress": "进行中", "completed": "已完成", "cancelled": "已取消"}.get(task.status, task.status),
            "截止时间": task.deadline.strftime("%Y-%m-%d %H:%M:%S") if task.deadline else "-",
            "创建时间": task.created_at.strftime("%Y-%m-%d %H:%M:%S") if task.created_at else "-",
            "完成时间": task.completed_at.strftime("%Y-%m-%d %H:%M:%S") if task.completed_at else "-"
        })
    
    return data, headers
